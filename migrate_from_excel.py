"""Migración de un solo uso: lee (solo lectura) el Excel de la porra y puebla
porra.db. No modifica el .xlsm ni nada dentro de agente_porra/.

Uso:
    python migrate_from_excel.py [ruta_al_xlsm]

Si no se pasa ruta, usa la del Excel de esta temporada un nivel por encima
de esta carpeta (../Liga BBVA 2026-2027 Apertura.xlsm).
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

from app.db import SessionLocal, init_db
from app.models import Jornada, Match, Participant, ParticipantTeam, Team

DEFAULT_XLSM = Path(__file__).resolve().parent.parent / "Liga BBVA 2026-2027 Apertura.xlsm"

# DatuBase!G2:G21
TEAMS_COL = 7  # G
TEAMS_FIRST_ROW = 2

# KinielaGuztiak: bloques de 11 filas desde la fila 3.
PARTICIPANT_FIRST_ROW = 3
PARTICIPANT_BLOCK_SIZE = 11
PARTICIPANT_NAME_COL = 2  # B
PARTICIPANT_TEAM_COL = 3  # C
PARTICIPANT_TEAMS_PER_BLOCK = 8

# PartiduenEmaitzak: bloques de 13 filas desde la fila 4 (cabecera en fila 3).
JORNADA_FIRST_HEADER_ROW = 3
JORNADA_BLOCK_SIZE = 13
JORNADA_MATCHES_PER_BLOCK = 10
MAX_JORNADAS = 38
COL_HOME = 2  # B
COL_AWAY = 3  # C
COL_HGOALS = 4  # D
COL_AGOALS = 6  # F

# KinielaSaikapena!W4:W... = jornadas trampa
TRAP_COL = 23  # W
TRAP_FIRST_ROW = 4


def _goals(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, str) and value.strip() in ("", "-"):
        return None
    return int(value)


def migrate(xlsm_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)

    init_db()
    session = SessionLocal()

    try:
        # --- Equipos ---------------------------------------------------
        ws = wb["DatuBase"]
        teams_by_name: dict[str, Team] = {}
        row = TEAMS_FIRST_ROW
        while True:
            name = ws.cell(row=row, column=TEAMS_COL).value
            if not name:
                break
            team = Team(name=str(name).strip())
            session.add(team)
            teams_by_name[team.name] = team
            row += 1
        session.flush()
        print(f"Equipos: {len(teams_by_name)}")

        # --- Participantes -----------------------------------------------
        ws = wb["KinielaGuztiak"]
        n_participants = 0
        block_start = PARTICIPANT_FIRST_ROW
        while True:
            name = ws.cell(row=block_start, column=PARTICIPANT_NAME_COL).value
            if not name:
                break
            participant = Participant(name=str(name).strip())
            for i in range(PARTICIPANT_TEAMS_PER_BLOCK):
                team_name = ws.cell(row=block_start + i, column=PARTICIPANT_TEAM_COL).value
                team_name = str(team_name).strip()
                team = teams_by_name.get(team_name)
                if team is None:
                    raise ValueError(
                        f"Participante '{name}': el equipo '{team_name}' no está en DatuBase!G."
                    )
                participant.picks.append(ParticipantTeam(team=team))
            session.add(participant)
            n_participants += 1
            block_start += PARTICIPANT_BLOCK_SIZE
        session.flush()
        print(f"Participantes: {n_participants}")

        # --- Jornadas y partidos -------------------------------------------
        ws = wb["PartiduenEmaitzak"]
        n_jornadas = 0
        n_matches = 0
        for n in range(1, MAX_JORNADAS + 1):
            hrow = JORNADA_FIRST_HEADER_ROW + (n - 1) * JORNADA_BLOCK_SIZE
            first_home = ws.cell(row=hrow + 1, column=COL_HOME).value
            if not first_home:
                continue  # jornada sin construir todavía (p.ej. 20-38 de esta Apertura)

            jornada = Jornada(number=n)
            session.add(jornada)
            for i in range(JORNADA_MATCHES_PER_BLOCK):
                r = hrow + 1 + i
                home_name = ws.cell(row=r, column=COL_HOME).value
                away_name = ws.cell(row=r, column=COL_AWAY).value
                if not home_name or not away_name:
                    continue
                home_name = str(home_name).strip()
                away_name = str(away_name).strip()
                home_team = teams_by_name.get(home_name)
                away_team = teams_by_name.get(away_name)
                if home_team is None or away_team is None:
                    raise ValueError(f"J{n}: equipo desconocido en '{home_name}' - '{away_name}'.")
                match = Match(
                    jornada_number=n,
                    home_team_id=home_team.id,
                    away_team_id=away_team.id,
                    home_goals=_goals(ws.cell(row=r, column=COL_HGOALS).value),
                    away_goals=_goals(ws.cell(row=r, column=COL_AGOALS).value),
                )
                session.add(match)
                n_matches += 1
            n_jornadas += 1
        session.flush()
        print(f"Jornadas: {n_jornadas} ({n_matches} partidos)")

        # --- Jornadas trampa -------------------------------------------------
        ws = wb["KinielaSaikapena"]
        n_trap = 0
        row = TRAP_FIRST_ROW
        while True:
            value = ws.cell(row=row, column=TRAP_COL).value
            if value is None:
                break
            jornada = session.get(Jornada, int(value))
            if jornada is not None:
                jornada.is_trap = True
                n_trap += 1
            row += 1
        print(f"Jornadas trampa marcadas: {n_trap}")

        session.commit()
        print("\nMigración completada. Verifica los totales contra KinielaSaikapena en el Excel.")
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


if __name__ == "__main__":
    xlsm_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSM
    if not xlsm_path.exists():
        print(f"No existe el fichero: {xlsm_path}")
        sys.exit(1)
    migrate(xlsm_path)
